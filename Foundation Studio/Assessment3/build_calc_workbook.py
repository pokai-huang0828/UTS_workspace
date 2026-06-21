# -*- coding: utf-8 -*-
"""
建立 A3 計算工作檔(Excel,公式即時計算)。
資料來源:APRA 季度會員趨勢(Mar 2026)、APRA 年度績效統計(2024-25)、Yahoo Finance、nib 年報。
所有衍生數字(覆蓋率、占比、市佔、利潤率、成長率、回歸)以 Excel 公式呈現,開啟時由 Excel 計算。
"""
import json
import numpy as np
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = Path(r"D:\KennyCode\UTS_workspace\Foundation Studio\Assessment3")
MT = BASE / "data" / "APRA_Membership_Trends_Mar2026.xlsx"
PF = BASE / "data" / "APRA_Performance_2024-25.xlsx"
OUT = BASE / "A3_計算工作檔.xlsx"
MONTH = {"Mar": 3, "Jun": 6, "Sep": 9, "Dec": 12}
QLAB = {3: "Q1", 6: "Q2", 9: "Q3", 12: "Q4"}

# ---------- styles ----------
NAVY = "1F3A5F"; LBLUE = "DCE6F1"; RED = "C0392B"; GREY = "F2F4F7"; YEL = "FFF3B0"
hfont = Font(bold=True, color="FFFFFF", size=11)
hfill = PatternFill("solid", fgColor=NAVY)
sub = Font(bold=True, color=NAVY, size=11)
note = Font(italic=True, color="666666", size=9)
nibfill = PatternFill("solid", fgColor="FBE4E0")
keyfill = PatternFill("solid", fgColor=YEL)
boxfill = PatternFill("solid", fgColor=GREY)
center = Alignment(horizontal="center", vertical="center")
left = Alignment(horizontal="left", vertical="center", wrap_text=True)
thin = Side(style="thin", color="C9D2DD")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr(ws, row, headers, start=1):
    for j, h in enumerate(headers):
        c = ws.cell(row=row, column=start + j, value=h)
        c.font = hfont; c.fill = hfill; c.alignment = center; c.border = border

def title(ws, text, src):
    ws["A1"] = text; ws["A1"].font = Font(bold=True, size=14, color=NAVY)
    ws["A2"] = src; ws["A2"].font = note

# =========================================================
# LOAD SOURCE DATA
# =========================================================
raw = pd.read_excel(MT, sheet_name="MembershipByAgeData", header=None, usecols=[0, 1, 2, 3, 4, 5])
raw.columns = ["Year", "MonthEnd", "State", "Gender", "Age", "Insured"]
for c in ["Year", "Age", "Insured"]:
    raw[c] = pd.to_numeric(raw[c], errors="coerce")
raw["MonthEnd"] = raw["MonthEnd"].astype(str).str.strip()
df = raw[(raw.Year.between(2007, 2026)) & raw.Age.notna() & raw.Insured.notna() & raw.MonthEnd.isin(MONTH)].copy()
df["date"] = pd.to_datetime(dict(year=df.Year.astype(int), month=df.MonthEnd.map(MONTH), day=1))
nat = df.groupby(["date", "Age"], as_index=False).Insured.sum()
tot = nat.groupby("date", as_index=False).Insured.sum().rename(columns={"Insured": "Total"})
u40 = nat[nat.Age < 40].groupby("date", as_index=False).Insured.sum().rename(columns={"Insured": "U40"})

pr = pd.read_excel(MT, sheet_name="MembershipData", header=None, usecols=[0, 1, 2, 3])
pr.columns = ["Year", "MonthEnd", "State", "Pop"]
pr["Year"] = pd.to_numeric(pr.Year, errors="coerce"); pr["Pop"] = pd.to_numeric(pr.Pop, errors="coerce")
pr["MonthEnd"] = pr.MonthEnd.astype(str).str.strip()
pop = pr[(pr.Year.between(2007, 2026)) & pr.Pop.notna() & pr.MonthEnd.isin(MONTH)].copy()
pop["date"] = pd.to_datetime(dict(year=pop.Year.astype(int), month=pop.MonthEnd.map(MONTH), day=1))
natpop = pop.groupby("date", as_index=False).Pop.sum()

cov = tot.merge(natpop, on="date").merge(u40, on="date").sort_values("date").reset_index(drop=True)
cov["q"] = cov.date.dt.month.map(QLAB)
cov["label"] = cov.date.dt.year.astype(str) + " " + cov.q

band = nat[nat.Age.isin([25, 30])].pivot(index="date", columns="Age", values="Insured")
band.columns = ["t25", "c30"]; band = band.sort_index().reset_index()
band["q"] = band.date.dt.month.map(QLAB)
band["label"] = band.date.dt.year.astype(str) + " " + band.q

# performance
db = pd.read_excel(PF, sheet_name="Database"); db["Value"] = pd.to_numeric(db["Value"], errors="coerce")
STATES = {"NSW", "VIC", "QLD", "SA", "WA", "TAS", "NT", "ACT"}
def item(g, name, hib=True):
    m = g["Data item"] == name
    if hib:
        m &= g["Business type"] == "Health insurance business"
    s_ = g.loc[m]; bs = s_[s_["State and territory"].isin(STATES)]
    return bs["Value"].sum() if len(bs) else s_["Value"].sum()
rows = []
for ent, g in db.groupby("Entity name"):
    prem = item(g, "Premium revenue"); claims = item(g, "Insurance claims")
    opex = item(g, "Other business expenses (inclusive of claims handling expenses)")
    pat = item(g, "Profit (loss) from continuing operations after income tax", hib=False)
    if prem and prem > 0:
        rows.append((ent, prem, claims, opex, pat))
perf = pd.DataFrame(rows, columns=["Entity", "Prem", "Claims", "Opex", "PAT"]).sort_values("Prem", ascending=False).reset_index(drop=True)
short = {"NIB Health Funds Ltd": "nib", "Medibank Private Limited": "Medibank", "BUPA HI Pty Ltd": "Bupa",
         "The Hospitals Contribution Fund of Australia Ltd": "HCF", "HBF Health Limited": "HBF",
         "Australian Unity Health Limited": "Australian Unity", "GMHBA Limited": "GMHBA",
         "Defence Health Limited": "Defence Health", "Teachers Federation Health Ltd": "Teachers Health",
         "CBHS Health Fund Limited": "CBHS", "Health Partners Limited": "Health Partners"}
perf["Short"] = perf.Entity.map(short).fillna(perf.Entity)

# share price
def load(name):
    j = json.loads((BASE / "data" / f"yahoo_{name}.json").read_text(encoding="utf-8"))
    r = j["chart"]["result"][0]; ts = pd.to_datetime(r["timestamp"], unit="s")
    return pd.Series(r["indicators"]["quote"][0]["close"], index=ts, name=name).dropna()
nhf, axjo = load("NHF"), load("AXJO")
sp = pd.concat([nhf, axjo], axis=1).dropna()
sp.index = sp.index.to_period("M").to_timestamp()
sp = sp.reset_index().rename(columns={"index": "date"})
sp["label"] = sp.date.dt.strftime("%Y-%m")

# uop
UOP = [("FY21", 204.9, 160.5), ("FY22", 237.0, None), ("FY23", 263.2, 191.1),
       ("FY24", 257.5, 181.6), ("FY25", 239.2, 198.6)]

# seasonality
seas = tot[tot.date >= pd.Timestamp("2010-01-01")].copy().sort_values("date").reset_index(drop=True)
seas["q"] = seas.date.dt.month.map(QLAB)
seas["label"] = seas.date.dt.year.astype(str) + " " + seas.q

# =========================================================
# BUILD WORKBOOK
# =========================================================
wb = Workbook()

# ---------- 說明 ----------
ws = wb.active; ws.title = "說明"
ws.column_dimensions["A"].width = 22; ws.column_dimensions["B"].width = 95
ws["A1"] = "評估三 計算工作檔 — NIB 案例"; ws["A1"].font = Font(bold=True, size=15, color=NAVY)
ws["A2"] = "口試後備檔:所有衍生數字都用 Excel 公式即時算,點任一格可看公式。圖在報告裡,計算過程在這裡。"
ws["A2"].font = note
r = 4
def line(k, v, kf=sub, vf=None):
    global r
    ws.cell(row=r, column=1, value=k).font = kf
    c = ws.cell(row=r, column=2, value=v); c.font = vf or Font(size=11); c.alignment = left
    r += 1
line("資料來源", "")
line("　會員/年齡/人口", "APRA《Quarterly Private Health Insurance Statistics — Membership Trends》, Mar 2026", note)
line("　市佔/利潤率", "APRA《Annual Private Health Insurance Performance Statistics》, FY2024-25(Health insurance business)", note)
line("　股價", "Yahoo Finance(NHF 與 ^AXJO 月收盤價)", note)
line("　UOP/NPAT", "nib Holdings 年報(FY21 五年摘要、FY23/FY25 業績公告)", note)
r += 1
line("分頁導覽", "")
guide = [
    ("1_覆蓋率", "住院險覆蓋率 = 投保總人數 ÷ 全國人口(對應報告圖一)"),
    ("2_年齡結構", "40 歲以下投保人占比(對應報告圖二:48.2% → 45.1%)"),
    ("3_市佔與利潤率", "各保險商保費市佔、淨利率、管理費用率(報告圖四、圖五)"),
    ("4_股價", "NHF vs ASX200 指數化(2011=100;報告圖三)"),
    ("5_獲利UOP", "nib 承保營運利潤 FY21-25(報告圖五)"),
    ("6_DiD", "差異中之差異 + 平行趨勢檢定(報告圖七、Part 3 核心)"),
    ("7_季節性回歸", "會員成長率對季別的多元回歸(R²、p 值;Part 3 鋪陳)"),
]
for k, v in guide:
    ws.cell(row=r, column=1, value=k).font = Font(bold=True, color=RED, size=11)
    ws.cell(row=r, column=2, value=v).alignment = left
    r += 1
r += 1
line("口試對應(老師會問)", "")
oral = [
    ("資料怎麼找", "→ 看「1_覆蓋率」「2_年齡結構」:從 APRA 季度表 MembershipByAgeData / MembershipData 擷取,自己彙總。"),
    ("用 Excel 還是別的", "→ 就是這個檔。每格都是 Excel 公式(SUM、除法、SLOPE、LINEST),非貼上的死值。"),
    ("單回歸 vs 多元回歸", "→「7_季節性回歸」是多元回歸(成長率 ~ Q2+Q3+Q4 三個虛擬變數);「6_DiD」是含交互項的回歸。"),
    ("R² 和 p 值怎麼解讀", "→「7_季節性回歸」:R²≈0.09(幾乎不解釋)、各季 p>0.05(不顯著)→ 沒有季節性,變動由政策/COVID 驅動。"),
    ("漲價對需求的因果", "→「6_DiD」:樸素 DiD −4.1%,但平行趨勢不成立(政策前 25-29 每季 −1.51% vs 30-34 −0.71%)→ 不可當因果。"),
]
for k, v in oral:
    ws.cell(row=r, column=1, value=k).font = sub
    cc = ws.cell(row=r, column=2, value=v); cc.alignment = left
    r += 1

# ---------- helper to write a series sheet ----------
def freeze(ws, cell="A4"):
    ws.freeze_panes = cell

# ---------- 1_覆蓋率 ----------
ws = wb.create_sheet("1_覆蓋率")
title(ws, "住院險覆蓋率(投保人數 ÷ 全國人口)", "來源:APRA Quarterly Membership Trends, Mar 2026。覆蓋率欄為公式。")
hdr(ws, 4, ["季別", "投保總人數(人)", "全國人口(人)", "覆蓋率 %"])
r0 = 5
for i, row in cov.iterrows():
    rr = r0 + i
    ws.cell(row=rr, column=1, value=row["label"]).border = border
    ws.cell(row=rr, column=2, value=int(row["Total"])).border = border
    ws.cell(row=rr, column=3, value=int(row["Pop"])).border = border
    f = ws.cell(row=rr, column=4, value=f"=B{rr}/C{rr}*100"); f.border = border; f.number_format = "0.0"
    ws.cell(row=rr, column=2).number_format = "#,##0"; ws.cell(row=rr, column=3).number_format = "#,##0"
rN = r0 + len(cov) - 1
for col, w in zip("ABCD", [11, 17, 16, 10]):
    ws.column_dimensions[col].width = w
freeze(ws)
# summary box
b = rN + 2
ws.cell(row=b, column=1, value="重點").font = sub
for k, formula, fmt in [
    ("最早(%s)" % cov.label.iloc[0], f"=D{r0}", "0.0"),
    ("最高峰", f"=MAX(D{r0}:D{rN})", "0.0"),
    ("最低谷", f"=MIN(D{r0}:D{rN})", "0.0"),
    ("最新(%s)" % cov.label.iloc[-1], f"=D{rN}", "0.0"),
]:
    b += 1
    ws.cell(row=b, column=1, value=k).fill = boxfill
    c = ws.cell(row=b, column=2, value=formula); c.number_format = fmt; c.font = Font(bold=True, color=NAVY)

# ---------- 2_年齡結構 ----------
ws = wb.create_sheet("2_年齡結構")
title(ws, "40 歲以下投保人占比", "來源:APRA Membership Trends(MembershipByAgeData)。占比欄為公式。")
hdr(ws, 4, ["季別", "投保總人數", "40 歲以下人數", "<40 占比 %"])
r0 = 5
for i, row in cov.iterrows():
    rr = r0 + i
    ws.cell(row=rr, column=1, value=row["label"]).border = border
    ws.cell(row=rr, column=2, value=int(row["Total"])).border = border
    ws.cell(row=rr, column=3, value=int(row["U40"])).border = border
    f = ws.cell(row=rr, column=4, value=f"=C{rr}/B{rr}*100"); f.border = border; f.number_format = "0.0"
    ws.cell(row=rr, column=2).number_format = "#,##0"; ws.cell(row=rr, column=3).number_format = "#,##0"
rN = r0 + len(cov) - 1
for col, w in zip("ABCD", [11, 15, 16, 12]):
    ws.column_dimensions[col].width = w
freeze(ws)
b = rN + 2
ws.cell(row=b, column=1, value="最早 → 最新").font = sub
ws.cell(row=b + 1, column=1, value=cov.label.iloc[0]).fill = boxfill
c = ws.cell(row=b + 1, column=2, value=f"=D{r0}"); c.number_format = "0.0"; c.font = Font(bold=True, color=NAVY)
ws.cell(row=b + 2, column=1, value=cov.label.iloc[-1]).fill = boxfill
c = ws.cell(row=b + 2, column=2, value=f"=D{rN}"); c.number_format = "0.0"; c.font = Font(bold=True, color=NAVY)
ws.cell(row=b + 3, column=1, value="下降(百分點)").fill = boxfill
c = ws.cell(row=b + 3, column=2, value=f"=D{r0}-D{rN}"); c.number_format = "0.0"; c.font = Font(bold=True, color=RED)

# ---------- 3_市佔與利潤率 ----------
ws = wb.create_sheet("3_市佔與利潤率")
title(ws, "各保險商:保費市佔 + 利潤率(FY2024-25)",
      "來源:APRA Annual PHI Performance, FY2024-25(Health insurance business)。市佔/利潤率/費用率為公式。")
hdr(ws, 4, ["保險商", "保費收入 $", "理賠 $", "其他費用 $", "稅後淨利 $",
            "市佔 %", "淨利率 %", "管理費用率 %"])
r0 = 5
for i, row in perf.iterrows():
    rr = r0 + i
    ws.cell(row=rr, column=1, value=row["Short"]).border = border
    for cidx, val in zip([2, 3, 4, 5], [row.Prem, row.Claims, row.Opex, row.PAT]):
        c = ws.cell(row=rr, column=cidx, value=int(val) if pd.notna(val) else None)
        c.number_format = "#,##0"; c.border = border
    rN_tmp = r0 + len(perf) - 1
rN = r0 + len(perf) - 1
for i, row in perf.iterrows():
    rr = r0 + i
    c = ws.cell(row=rr, column=6, value=f"=B{rr}/SUM($B${r0}:$B${rN})*100"); c.number_format = "0.0"; c.border = border
    c = ws.cell(row=rr, column=7, value=f"=(1-(C{rr}+D{rr})/B{rr})*100"); c.number_format = "0.0"; c.border = border
    c = ws.cell(row=rr, column=8, value=f"=D{rr}/B{rr}*100"); c.number_format = "0.0"; c.border = border
    if row["Short"] == "nib":
        for cc in range(1, 9):
            ws.cell(row=rr, column=cc).fill = nibfill
for col, w in zip("ABCDEFGH", [16, 14, 13, 13, 13, 9, 11, 14]):
    ws.column_dimensions[col].width = w
freeze(ws)
b = rN + 2
ws.cell(row=b, column=1, value="產業總保費 $").fill = boxfill
c = ws.cell(row=b, column=2, value=f"=SUM(B{r0}:B{rN})"); c.number_format = "#,##0"; c.font = Font(bold=True, color=NAVY)
ws.cell(row=b, column=4, value="保險商家數").fill = boxfill
ws.cell(row=b, column=5, value=f"=COUNTA(A{r0}:A{rN})").font = Font(bold=True, color=NAVY)
ws.cell(row=b + 1, column=1, value="淨利率 = 1 −(理賠+其他費用)÷保費  |  市佔 = 該商保費 ÷ 全產業保費").font = note

# ---------- 4_股價 ----------
ws = wb.create_sheet("4_股價")
title(ws, "NHF(nib) vs ASX200 股價,指數化", "來源:Yahoo Finance 月收盤價。指數欄 = 當期 ÷ 首期 ×100(公式)。")
hdr(ws, 4, ["月份", "NHF 收盤", "ASX200 收盤", "NHF 指數", "ASX 指數"])
r0 = 5
for i, row in sp.iterrows():
    rr = r0 + i
    ws.cell(row=rr, column=1, value=row["label"]).border = border
    ws.cell(row=rr, column=2, value=round(float(row["NHF"]), 3)).border = border
    ws.cell(row=rr, column=3, value=round(float(row["AXJO"]), 1)).border = border
    c = ws.cell(row=rr, column=4, value=f"=B{rr}/$B${r0}*100"); c.number_format = "0"; c.border = border
    c = ws.cell(row=rr, column=5, value=f"=C{rr}/$C${r0}*100"); c.number_format = "0"; c.border = border
    ws.cell(row=rr, column=2).number_format = "0.00"; ws.cell(row=rr, column=3).number_format = "#,##0"
rN = r0 + len(sp) - 1
for col, w in zip("ABCDE", [10, 11, 13, 11, 11]):
    ws.column_dimensions[col].width = w
freeze(ws)
b = rN + 2
ws.cell(row=b, column=1, value="最新 NHF 指數(倍)").fill = boxfill
c = ws.cell(row=b, column=2, value=f"=D{rN}/100"); c.number_format = "0.0\"x\""; c.font = Font(bold=True, color=RED)
ws.cell(row=b + 1, column=1, value="最新 ASX 指數(倍)").fill = boxfill
c = ws.cell(row=b + 1, column=2, value=f"=E{rN}/100"); c.number_format = "0.0\"x\""; c.font = Font(bold=True, color=NAVY)

# ---------- 5_獲利UOP ----------
ws = wb.create_sheet("5_獲利UOP")
title(ws, "nib 承保營運利潤(UOP)與稅後淨利(NPAT)", "來源:nib Holdings 年報。YoY 為公式。")
hdr(ws, 4, ["財年", "UOP A$m", "NPAT A$m", "UOP YoY %"])
r0 = 5
for i, (fy, uop, npat) in enumerate(UOP):
    rr = r0 + i
    ws.cell(row=rr, column=1, value=fy).border = border
    ws.cell(row=rr, column=2, value=uop).border = border; ws.cell(row=rr, column=2).number_format = "0.0"
    ws.cell(row=rr, column=3, value=npat).border = border
    if npat is not None:
        ws.cell(row=rr, column=3).number_format = "0.0"
    if i > 0:
        c = ws.cell(row=rr, column=4, value=f"=(B{rr}-B{rr-1})/B{rr-1}*100"); c.number_format = "+0.0;-0.0"; c.border = border
rN = r0 + len(UOP) - 1
for col, w in zip("ABCD", [9, 11, 11, 11]):
    ws.column_dimensions[col].width = w
ws.cell(row=rN + 2, column=1, value="FY23 見頂、FY24-25 受理賠通膨回落。FY22 NPAT 未列。").font = note

# ---------- 6_DiD ----------
ws = wb.create_sheet("6_DiD")
title(ws, "差異中之差異(DiD)+ 平行趨勢檢定",
      "處理組 25-29(2019/4 起享年齡折扣)vs 對照組 30-34。來源:APRA Membership Trends。")
hdr(ws, 4, ["季別", "時間序", "處理 25-29", "對照 30-34", "ln(處理)", "ln(對照)", "期間"])
did = band[(band.date >= "2016-03-01") & (band.date <= "2019-12-01")].reset_index(drop=True)
r0 = 5
prerows = []; postrows = []; pretrend = []
for i, row in did.iterrows():
    rr = r0 + i
    ws.cell(row=rr, column=1, value=row["label"]).border = border
    ws.cell(row=rr, column=2, value=i + 1).border = border
    ws.cell(row=rr, column=3, value=int(row["t25"])).border = border; ws.cell(row=rr, column=3).number_format = "#,##0"
    ws.cell(row=rr, column=4, value=int(row["c30"])).border = border; ws.cell(row=rr, column=4).number_format = "#,##0"
    ws.cell(row=rr, column=5, value=f"=LN(C{rr})").border = border; ws.cell(row=rr, column=5).number_format = "0.0000"
    ws.cell(row=rr, column=6, value=f"=LN(D{rr})").border = border; ws.cell(row=rr, column=6).number_format = "0.0000"
    d = row["date"]
    period = ""
    if pd.Timestamp("2017-03-01") <= d <= pd.Timestamp("2019-03-01"):
        period = "pre"; prerows.append(rr)
    elif pd.Timestamp("2019-06-01") <= d <= pd.Timestamp("2019-12-01"):
        period = "post"; postrows.append(rr)
    if pd.Timestamp("2016-03-01") <= d <= pd.Timestamp("2019-03-01"):
        pretrend.append(rr)
    ws.cell(row=rr, column=7, value=period).border = border
    if period:
        ws.cell(row=rr, column=7).fill = PatternFill("solid", fgColor="E8EEF6")
rN = r0 + len(did) - 1
for col, w in zip("ABCDEFG", [10, 8, 12, 12, 10, 10, 8]):
    ws.column_dimensions[col].width = w
freeze(ws)

pre_s, pre_e = prerows[0], prerows[-1]
post_s, post_e = postrows[0], postrows[-1]
pt_s, pt_e = pretrend[0], pretrend[-1]
b = rN + 2
ws.cell(row=b, column=1, value="2×2 DiD(對 ln 投保人數取平均)").font = sub
tbl = [
    ("處理組 pre 均值", f"=AVERAGE(E{pre_s}:E{pre_e})"),
    ("處理組 post 均值", f"=AVERAGE(E{post_s}:E{post_e})"),
    ("對照組 pre 均值", f"=AVERAGE(F{pre_s}:F{pre_e})"),
    ("對照組 post 均值", f"=AVERAGE(F{post_s}:F{post_e})"),
]
for k, f in tbl:
    b += 1
    ws.cell(row=b, column=1, value=k).fill = boxfill
    c = ws.cell(row=b, column=2, value=f); c.number_format = "0.0000"
dlt = b + 1; ws.cell(row=dlt, column=1, value="Δln 處理(post-pre)").fill = boxfill
ws.cell(row=dlt, column=2, value=f"=B{b-2}-B{b-3}").number_format = "+0.0000;-0.0000"
dlc = b + 2; ws.cell(row=dlc, column=1, value="Δln 對照(post-pre)").fill = boxfill
ws.cell(row=dlc, column=2, value=f"=B{b}-B{b-1}").number_format = "+0.0000;-0.0000"
ddr = b + 3
ws.cell(row=ddr, column=1, value="DiD = Δ處理 − Δ對照").fill = keyfill
c = ws.cell(row=ddr, column=2, value=f"=B{dlt}-B{dlc}"); c.number_format = "+0.0000;-0.0000"; c.font = Font(bold=True, color=RED)
ws.cell(row=ddr, column=3, value="≈ %").fill = keyfill
c = ws.cell(row=ddr, column=4, value=f"=B{ddr}*100"); c.number_format = "+0.0\"%\";-0.0\"%\""; c.font = Font(bold=True, color=RED)

pb = ddr + 2
ws.cell(row=pb, column=1, value="平行趨勢檢定(政策前 2016Q1-2019Q1,ln 對時間序回歸的斜率)").font = sub
ws.cell(row=pb + 1, column=1, value="處理 25-29 每季成長 %").fill = boxfill
c = ws.cell(row=pb + 1, column=2, value=f"=SLOPE(E{pt_s}:E{pt_e},B{pt_s}:B{pt_e})*100"); c.number_format = "+0.00;-0.00"; c.font = Font(bold=True, color=RED)
ws.cell(row=pb + 2, column=1, value="對照 30-34 每季成長 %").fill = boxfill
c = ws.cell(row=pb + 2, column=2, value=f"=SLOPE(F{pt_s}:F{pt_e},B{pt_s}:B{pt_e})*100"); c.number_format = "+0.00;-0.00"; c.font = Font(bold=True, color=NAVY)
ws.cell(row=pb + 3, column=1, value="差距(百分點/季)").fill = boxfill
c = ws.cell(row=pb + 3, column=2, value=f"=B{pb+1}-B{pb+2}"); c.number_format = "+0.00;-0.00"; c.font = Font(bold=True, color=RED)
ws.cell(row=pb + 4, column=1, value="差距大 → 平行趨勢不成立 → 上面的 DiD 不能解讀為因果。").font = note

# ---------- 7_季節性回歸 ----------
ws = wb.create_sheet("7_季節性回歸")
title(ws, "會員成長率的季節性:多元回歸(成長率 ~ Q2 + Q3 + Q4)",
      "來源:APRA Membership Trends 總投保人數。成長率/虛擬變數/回歸統計皆為公式。")
hdr(ws, 4, ["季別", "時間序", "投保總人數", "QoQ 成長 %", "Q2", "Q3", "Q4"])
r0 = 5
for i, row in seas.iterrows():
    rr = r0 + i
    ws.cell(row=rr, column=1, value=row["label"]).border = border
    ws.cell(row=rr, column=2, value=i + 1).border = border
    ws.cell(row=rr, column=3, value=int(row["Total"])).border = border; ws.cell(row=rr, column=3).number_format = "#,##0"
    if i > 0:
        c = ws.cell(row=rr, column=4, value=f"=(C{rr}-C{rr-1})/C{rr-1}*100"); c.number_format = "+0.00;-0.00"; c.border = border
    ws.cell(row=rr, column=5, value=f'=IF(A{rr}="","",IF(RIGHT(A{rr},2)="Q2",1,0))').border = border
    ws.cell(row=rr, column=6, value=f'=IF(RIGHT(A{rr},2)="Q3",1,0)').border = border
    ws.cell(row=rr, column=7, value=f'=IF(RIGHT(A{rr},2)="Q4",1,0)').border = border
rN = r0 + len(seas) - 1
yr0 = r0 + 1  # first row with growth (2010 Q2)
for col, w in zip("ABCDEFG", [10, 8, 14, 12, 6, 6, 6]):
    ws.column_dimensions[col].width = w
freeze(ws)

# regression output block (to the right)
oc = 9  # column I
def put(rrow, ccol, val, fmt=None, font=None, fill=None):
    c = ws.cell(row=rrow, column=ccol, value=val)
    if fmt: c.number_format = fmt
    if font: c.font = font
    if fill: c.fill = fill
    return c
yrng = f"D{yr0}:D{rN}"; xrng = f"E{yr0}:G{rN}"
LIN = f"LINEST({yrng},{xrng},TRUE,TRUE)"
put(4, oc, "回歸輸出(LINEST)", font=sub)
labels = [
    ("樣本數 n", f"=COUNT({yrng})", "0"),
    ("基準 = Q1 平均成長 %", f"=INDEX({LIN},1,4)", "+0.000;-0.000"),
    ("Q2 係數", f"=INDEX({LIN},1,3)", "+0.000;-0.000"),
    ("Q3 係數", f"=INDEX({LIN},1,2)", "+0.000;-0.000"),
    ("Q4 係數", f"=INDEX({LIN},1,1)", "+0.000;-0.000"),
    ("Q2 標準誤", f"=INDEX({LIN},2,3)", "0.000"),
    ("Q3 標準誤", f"=INDEX({LIN},2,2)", "0.000"),
    ("Q4 標準誤", f"=INDEX({LIN},2,1)", "0.000"),
    ("R²", f"=INDEX({LIN},3,1)", "0.000"),
    ("F 統計量", f"=INDEX({LIN},4,1)", "0.000"),
    ("殘差自由度 df", f"=INDEX({LIN},4,2)", "0"),
]
rr = 5
cellref = {}
for name, formula, fmt in labels:
    put(rr, oc, name, fill=boxfill)
    c = put(rr, oc + 1, formula, fmt, Font(bold=True, color=NAVY))
    cellref[name] = f"{get_column_letter(oc+1)}{rr}"
    rr += 1
# p-values from coef/se and df
put(rr, oc, "Q2 p 值", fill=boxfill)
put(rr, oc + 1, f"=TDIST(ABS({cellref['Q2 係數']}/{cellref['Q2 標準誤']}),{cellref['殘差自由度 df']},2)", "0.000", Font(bold=True, color=RED)); rr += 1
put(rr, oc, "Q3 p 值", fill=boxfill)
put(rr, oc + 1, f"=TDIST(ABS({cellref['Q3 係數']}/{cellref['Q3 標準誤']}),{cellref['殘差自由度 df']},2)", "0.000", Font(bold=True, color=RED)); rr += 1
put(rr, oc, "Q4 p 值", fill=boxfill)
put(rr, oc + 1, f"=TDIST(ABS({cellref['Q4 係數']}/{cellref['Q4 標準誤']}),{cellref['殘差自由度 df']},2)", "0.000", Font(bold=True, color=RED)); rr += 1
put(rr, oc, "整體 F 檢定 p 值", fill=boxfill)
put(rr, oc + 1, f"=FDIST({cellref['F 統計量']},3,{cellref['殘差自由度 df']})", "0.000", Font(bold=True, color=RED)); rr += 1
rr += 1
put(rr, oc, "簡單線性趨勢線(對時間序)", font=sub); rr += 1
put(rr, oc, "R²(趨勢線)", fill=boxfill)
put(rr, oc + 1, f"=RSQ({yrng},B{yr0}:B{rN})", "0.0000", Font(bold=True, color=NAVY)); rr += 1
put(rr, oc, "成長率平均 %", fill=boxfill)
put(rr, oc + 1, f"=AVERAGE({yrng})", "0.00", Font(bold=True, color=NAVY)); rr += 1
put(rr, oc, "成長率標準差 %", fill=boxfill)
put(rr, oc + 1, f"=STDEV({yrng})", "0.00", Font(bold=True, color=NAVY)); rr += 1
rr += 1
concl = ("解讀:R²≈0.09,各季 p 值均 > 0.05、整體 F 檢定也不顯著 → 在 5% 水準下沒有證據顯示成長率有季節性。"
         "趨勢線 R²≈0 → 也不是單向遞增/遞減。會員變動由政策與 COVID 等一次性事件驅動,"
         "因此單純相關會誤導,Part 3 才需要準實驗(DiD)來識別因果。")
cc = ws.cell(row=rr, column=oc, value=concl); cc.font = note; cc.alignment = left
ws.merge_cells(start_row=rr, start_column=oc, end_row=rr + 4, end_column=oc + 5)
for col in ["I", "J", "K", "L", "M", "N"]:
    ws.column_dimensions[col].width = 16

wb.save(OUT)
print("SAVED:", OUT)
print("Sheets:", wb.sheetnames)
