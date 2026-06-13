# -*- coding: utf-8 -*-
"""List APRA workbook structure: sheet names + dimensions + first-cell preview."""
import openpyxl

PATH = r"D:\KennyWorkLife\UTS_workspace\Foundation Studio\Assessment3\data\APRA_Membership_Trends_Mar2026.xlsx"

wb = openpyxl.load_workbook(PATH, read_only=True, data_only=True)
print("TOTAL SHEETS:", len(wb.sheetnames))
print("=" * 70)
for name in wb.sheetnames:
    ws = wb[name]
    # grab the top-left 3x1 to hint at what the sheet is about
    first = ""
    for row in ws.iter_rows(min_row=1, max_row=4, max_col=2, values_only=True):
        vals = [str(c) for c in row if c is not None]
        if vals:
            first = " | ".join(vals)[:90]
            break
    print(f"[{name}]  rows={ws.max_row} cols={ws.max_column}")
    print(f"     ^ {first}")
wb.close()
